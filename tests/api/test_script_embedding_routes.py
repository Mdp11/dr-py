"""M2/M3 embedded-evaluation route tests (TrustedRunner injected). Route-level
coverage lands in Tasks 11-13; this file starts with the script_eval helper
and now covers Task 11's `POST /tables/evaluate` script-column wiring."""

from __future__ import annotations

import io
from collections.abc import Iterator

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from data_rover.api.main import create_app
from data_rover.api.script_eval import close_script_context, open_script_context
from data_rover.api.script_runner import get_runner
from data_rover.api.settings import Settings
from data_rover.api.snippet_concurrency import ConcurrencyGuard, concurrency_guard
from data_rover.core.metamodel.schema import ElementType, Metamodel
from data_rover.core.model.model import Model
from tests.script.trusted_runner import TrustedRunner

from .conftest import AUTH_HEADERS, papi, seed_default_project


def _settings(**kw) -> Settings:
    return Settings(dev_seed=True, **kw)


@pytest.fixture(autouse=True)
def _reset_concurrency_guard():
    """Safety net for the module-singleton `concurrency_guard`: the tests in
    this file release every slot they acquire (matching the brief's
    `close_script_context` calls), but assert that held here so a leaked
    acquire fails fast in THIS test rather than silently starving a later
    one."""
    yield
    assert concurrency_guard._global_count == 0
    concurrency_guard._per_user_count.clear()


@pytest.fixture(scope="session")
def small_model() -> Model:
    """A minimal Model (one `Building` element) for exercising
    `open_script_context`/`ScriptEvalContext` without a full session fixture."""
    metamodel = Metamodel(elements=[ElementType(name="Building")])
    model = Model(metamodel)
    model.create_element("Building")
    return model


def test_guard_global_slot() -> None:
    g = ConcurrencyGuard()
    assert g.try_acquire_global(global_limit=1)
    assert not g.try_acquire_global(global_limit=1)
    g.release_global()
    assert g.try_acquire_global(global_limit=1)
    g.release_global()


def test_open_context_modes(small_model) -> None:
    s = _settings()
    ctx, acquired = open_script_context(None, None, s, needs_script=False)
    assert ctx is None and not acquired

    ctx, acquired = open_script_context(None, None, s, needs_script=True)
    assert ctx is not None and not acquired          # unavailable mode
    res = ctx.call("def value(els): return 1", "value", ["x"])
    assert res.error is not None and res.error.kind == "unavailable"
    close_script_context(ctx, acquired)

    runner = TrustedRunner()
    ctx, acquired = open_script_context(runner, small_model, s, needs_script=True)
    assert ctx is not None and acquired
    close_script_context(ctx, acquired)


def test_open_context_busy(small_model) -> None:
    s = _settings(snippet_concurrency=1)
    runner = TrustedRunner()
    ctx1, a1 = open_script_context(runner, small_model, s, needs_script=True)
    ctx2, a2 = open_script_context(runner, small_model, s, needs_script=True)
    assert a1 and not a2
    assert ctx2 is not None
    res = ctx2.call("def value(els): return 1", "value", ["x"])
    assert res.error is not None and "busy" in res.error.message
    close_script_context(ctx2, a2)
    close_script_context(ctx1, a1)
    # slot actually freed:
    assert concurrency_guard.try_acquire_global(global_limit=1)
    concurrency_guard.release_global()


# ---------------------------------------------------------------------------
# Task 11: POST /tables/evaluate script-column wiring
# ---------------------------------------------------------------------------

THING_MM = """
elements:
  - name: Thing
    key: [name]
    properties:
      - {name: name, datatype: string, multiplicity: "1"}
"""


@pytest.fixture
def app() -> Iterator[FastAPI]:
    """Separate fixture (rather than building it inside `client`) so tests
    that need to swap `get_runner`'s override mid-test (the
    runner-unavailable case) can reach `app.dependency_overrides` with a
    properly-typed `FastAPI` object — mirrors
    `tests/api/test_snippets_routes.py::app`."""
    seed_default_project()
    application = create_app()
    application.dependency_overrides[get_runner] = lambda: TrustedRunner()
    yield application
    application.dependency_overrides.clear()


@pytest.fixture
def client(app: FastAPI) -> TestClient:
    c = TestClient(app)
    c.headers.update(AUTH_HEADERS)
    return c


@pytest.fixture
def seed_thing_model(client: TestClient) -> None:
    """Metamodel declaring `Thing` (a `name` string property) plus a few
    `Thing` elements, seeded into the default session via the HTTP routes
    (mirrors `tests/api/test_ops_route.py`'s `client` fixture shape). Includes
    a `name == "B"` element so tests exercising a conditional-raise snippet
    (`if els[0].name == 'B': raise ...`) get both an erroring row and clean
    rows in the same table."""
    r = client.post(
        papi("/metamodel"),
        content=THING_MM,
        headers={"content-type": "application/x-yaml"},
    )
    assert r.status_code == 200, r.text
    r = client.post(
        papi("/model"),
        json={
            "elements": [
                {"id": "t1", "type_name": "Thing", "properties": {"name": "Alpha"}},
                {"id": "t2", "type_name": "Thing", "properties": {"name": "Beta"}},
                {"id": "t3", "type_name": "Thing", "properties": {"name": "B"}},
            ],
            "relationships": [],
        },
    )
    assert r.status_code == 200, r.text


def _script_table(code: str) -> dict:
    return {
        "row_source": {"kind": "scope", "types": ["Thing"]},
        "columns": [{"kind": "script", "snippet": {"definition": {"code": code}}}],
    }


def test_evaluate_script_column_end_to_end(
    client: TestClient, seed_thing_model: None
) -> None:
    r = client.post(
        papi("/tables/evaluate"),
        json={"definition": _script_table("def value(els): return els[0].name")},
        headers=AUTH_HEADERS,
    )
    assert r.status_code == 200, r.text
    body = r.json()
    kinds = {c["kind"] for row in body["rows"] for c in row["cells"]}
    assert kinds == {"value"}
    assert body["warnings"] == []


def test_evaluate_script_error_cell_and_no_cache_poisoning(
    client: TestClient, seed_thing_model: None
) -> None:
    defn = _script_table("def value(els): raise RuntimeError('boom')")
    r = client.post(papi("/tables/evaluate"), json={"definition": defn}, headers=AUTH_HEADERS)
    assert r.status_code == 200
    cell = r.json()["rows"][0]["cells"][0]
    assert cell["kind"] == "error" and "boom" in cell["message"]
    # errored evaluations must NOT be served from the order cache: fix the
    # snippet inside the same rev and re-evaluate — cells recompute (a poisoned
    # order cache would be invisible here, so assert on the cache directly):
    from data_rover.api.session import get_session

    assert len(get_session().table_order_cache._d) == 0


def test_evaluate_runner_unavailable_degrades(
    client: TestClient, app: FastAPI, seed_thing_model: None
) -> None:
    app.dependency_overrides[get_runner] = lambda: None
    r = client.post(
        papi("/tables/evaluate"),
        json={"definition": _script_table("def value(els): return 1")},
        headers=AUTH_HEADERS,
    )
    assert r.status_code == 200
    cell = r.json()["rows"][0]["cells"][0]
    assert cell["kind"] == "error" and "unavailable" in cell["message"]


def test_evaluate_dangling_snippet_ref(
    client: TestClient, seed_thing_model: None
) -> None:
    defn = {
        "row_source": {"kind": "scope", "types": ["Thing"]},
        "columns": [{"kind": "script", "snippet": {"ref": "no-such-artifact"}}],
    }
    r = client.post(papi("/tables/evaluate"), json={"definition": defn}, headers=AUTH_HEADERS)
    assert r.status_code == 200
    cell = r.json()["rows"][0]["cells"][0]
    assert cell["kind"] == "error" and "not found" in cell["message"]


def test_evaluate_saved_snippet_ref_and_fingerprint(
    client: TestClient, seed_thing_model: None
) -> None:
    # create a snippet artifact, reference it, evaluate; then edit it and
    # confirm the next evaluation reflects the new code (fingerprint moved).
    r = client.post(
        papi("/artifacts"),
        json={
            "kind": "code_snippet",
            "name": "col",
            "payload": {"code": "def value(els): return 'v1'"},
        },
        headers=AUTH_HEADERS,
    )
    assert r.status_code == 201, r.text
    art = r.json()
    defn = {
        "row_source": {"kind": "scope", "types": ["Thing"]},
        "columns": [{"kind": "script", "snippet": {"ref": art["id"]}}],
    }
    r = client.post(papi("/tables/evaluate"), json={"definition": defn}, headers=AUTH_HEADERS)
    assert r.json()["rows"][0]["cells"][0]["value"] == "v1"
    r = client.put(
        papi(f"/artifacts/{art['id']}"),
        json={
            "artifact_rev": art["artifact_rev"],
            "payload": {"code": "def value(els): return 'v2'"},
        },
        headers=AUTH_HEADERS,
    )
    assert r.status_code == 200, r.text
    r = client.post(papi("/tables/evaluate"), json={"definition": defn}, headers=AUTH_HEADERS)
    assert r.json()["rows"][0]["cells"][0]["value"] == "v2"


# ---------------------------------------------------------------------------
# Task 12: POST /tables/export script-column wiring
# ---------------------------------------------------------------------------


def test_export_script_column_with_errors_gets_marker_and_notice(
    client: TestClient, seed_thing_model: None
) -> None:
    defn = _script_table(
        "def value(els):\n"
        "    if els[0].name == 'B': raise RuntimeError('boom')\n"
        "    return els[0].name"
    )
    r = client.post(papi("/tables/export"), json={"definition": defn}, headers=AUTH_HEADERS)
    assert r.status_code == 200, r.text
    assert r.headers.get("X-Table-Script-Errors") == "true"
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    ws = wb.active
    assert ws is not None
    texts = [str(row[0].value) for row in ws.iter_rows() if row and row[0].value is not None]
    assert any(t.startswith("#ERROR:") for t in texts)
    assert "script" in texts[-1].lower()  # trailing notice row


def test_export_script_column_no_errors_no_notice(
    client: TestClient, seed_thing_model: None
) -> None:
    defn = _script_table("def value(els): return els[0].name")
    r = client.post(papi("/tables/export"), json={"definition": defn}, headers=AUTH_HEADERS)
    assert r.status_code == 200, r.text
    assert "X-Table-Script-Errors" not in r.headers
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    ws = wb.active
    assert ws is not None
    texts = [str(row[0].value) for row in ws.iter_rows() if row and row[0].value is not None]
    assert not any(t.startswith("#ERROR:") for t in texts)
