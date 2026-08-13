"""POST /tables/export: whole-table xlsx export (Task 10)."""

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from data_rover.api.main import create_app
from data_rover.api.table_export import _sheet_title

from .conftest import AUTH_HEADERS, papi, seed_default_project
from .test_artifacts_routes import _bootstrap_model


@pytest.fixture
def client() -> TestClient:
    seed_default_project()
    c = TestClient(create_app())
    c.headers.update(AUTH_HEADERS)
    return c


def test_export_xlsx_has_header_and_rows(client):
    _bootstrap_model(client)
    body = {
        "definition": {
            "row_source": {"kind": "scope", "types": ["Block"]},
            "columns": [
                {"kind": "element", "source": {"kind": "row"}, "header": "Block"},
                {
                    "kind": "property",
                    "source": {"kind": "row"},
                    "name": "mass",
                    "header": "Mass",
                },
            ],
        }
    }
    r = client.post(papi("/tables/export"), json=body, headers=AUTH_HEADERS)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws is not None
    assert [c.value for c in ws[1]] == ["Block", "Mass"]  # header row
    assert ws.max_row >= 2


def test_export_truncation_header(client):
    _bootstrap_model(client)
    # not asserting the exact flag here unless the fixture exceeds max_rows;
    # assert the header key is absent for a small table
    body = {
        "definition": {
            "row_source": {"kind": "scope", "types": ["Block"]},
            "columns": [{"kind": "element", "source": {"kind": "row"}}],
        }
    }
    r = client.post(papi("/tables/export"), json=body, headers=AUTH_HEADERS)
    assert "x-table-truncated" not in {k.lower() for k in r.headers}


def test_export_includes_full_navigation_cell_beyond_cell_cap(client):
    # Regression: export used min(cell_cap, max_cell_elements), so a
    # navigation column's per-column display cap silently truncated exported
    # cells. The workbook must carry the COMPLETE reached set.
    _bootstrap_model(client)
    body = {
        "definition": {
            "row_source": {"kind": "scope", "types": ["Block"]},
            "columns": [
                {"kind": "element", "source": {"kind": "row"}, "header": "Block"},
                {
                    "kind": "navigation",
                    "source": {"kind": "row"},
                    "mode": "collapse",
                    "cell_cap": 1,
                    "header": "Parts",
                    "navigation": {"definition": {
                        "kind": "path",
                        "start": {"kind": "row"},
                        "steps": [{"kind": "relationship",
                                   "relationship_type": "BlockHasPart",
                                   "direction": "out"}],
                    }},
                },
            ],
        }
    }
    r = client.post(papi("/tables/export"), json=body, headers=AUTH_HEADERS)
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws is not None
    root_cell = str(
        next(row[1].value for row in ws.iter_rows(min_row=2) if row[0].value == "root")
    )
    assert "p1" in root_cell and "p2" in root_cell  # both parts, despite cell_cap=1


def test_export_skips_hidden_columns(client):
    # The middle column (navigation, hidden) is still EVALUATED — the third
    # column references it by index and reads a property off the elements it
    # reaches — but it must be omitted from the exported header/body.
    _bootstrap_model(client)
    body = {
        "definition": {
            "row_source": {
                "kind": "scope",
                "types": ["Block"],
                "criteria": [
                    {"type": "name_id", "field": "name", "op": "equals", "value": "root"}
                ],
            },
            "columns": [
                {"kind": "element", "source": {"kind": "row"}, "header": "Block"},
                {
                    "kind": "navigation",
                    "source": {"kind": "row"},
                    "mode": "expand",
                    "hidden": True,
                    "header": "Navigation",
                    "navigation": {"definition": {
                        "kind": "path",
                        "start": {"kind": "row"},
                        "steps": [{"kind": "relationship",
                                   "relationship_type": "BlockHasPart",
                                   "direction": "out"}],
                    }},
                },
                {
                    "kind": "property",
                    "source": {"kind": "column", "index": 1},
                    "name": "mass",
                    "header": "Mass",
                },
            ],
        }
    }
    r = client.post(papi("/tables/export"), json=body, headers=AUTH_HEADERS)
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws is not None
    header = [c.value for c in ws[1]]
    assert len(header) == 2  # hidden column absent
    assert "Navigation" not in header
    assert header == ["Block", "Mass"]
    # the hidden navigation column still evaluated: two rows (p1, p2), and the
    # dependent visible property column correctly read mass off the reached
    # elements rather than blowing up or reading the wrong binding.
    body_rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    assert len(body_rows) == 2
    assert all(row == ["root", 1.0] for row in body_rows)


def test_export_styling_autofit_filters_borders(client):
    # Item 11: the workbook ships with header-filter dropdowns, borders, bold
    # header, frozen header row, and autofitted column widths.
    _bootstrap_model(client)
    body = {
        "definition": {
            "row_source": {"kind": "scope", "types": ["Block"]},
            "columns": [
                {"kind": "element", "source": {"kind": "row"}, "header": "Block"},
                {
                    "kind": "property",
                    "source": {"kind": "row"},
                    "name": "mass",
                    "header": "Mass",
                    # a definition width must NOT drive the export any more
                    "width_px": 700,
                },
            ],
        }
    }
    r = client.post(papi("/tables/export"), json=body, headers=AUTH_HEADERS)
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws is not None
    # header filters span the data range (header row through the last data row)
    assert ws.auto_filter.ref is not None
    assert ws.auto_filter.ref.startswith("A1:B")
    # frozen header row survives the library swap
    assert ws.freeze_panes == "A2"
    # bold header with a heavier bottom edge; thin borders on data cells
    hdr = ws["A1"]
    assert hdr.font.b
    assert hdr.border.bottom.style == "medium"
    data = ws["A2"]
    assert data.border.left.style == "thin"
    assert data.border.bottom.style == "thin"
    # autofit set a real width, and the 700px definition width did not win
    # (700px under the old px/7 heuristic would exceed 90 char-units)
    w = ws.column_dimensions["B"].width
    assert w is not None and 0 < w < 90


def _export_long_name_col_width(client: TestClient) -> float:
    """One Block whose name is 200 chars; export a one-column table and
    return column A's autofitted width in char units (~px/7)."""
    _bootstrap_model(client)
    client.post(
        papi("/model/elements"),
        json={"type": "Block", "properties": {"name": "x" * 200, "mass": 1.0}},
        headers=AUTH_HEADERS,
    )
    body = {
        "definition": {
            "row_source": {"kind": "scope", "types": ["Block"]},
            "columns": [{"kind": "element", "source": {"kind": "row"}, "header": "Block"}],
        }
    }
    r = client.post(papi("/tables/export"), json=body, headers=AUTH_HEADERS)
    assert r.status_code == 200, r.text
    ws = load_workbook(io.BytesIO(r.content)).active
    assert ws is not None
    w = ws.column_dimensions["A"].width
    assert w is not None
    return w


def test_autofit_cap_default_allows_wider_columns(client):
    # U-2: the cap default moved from 300px (~43 chars — the ceiling the owner
    # kept hitting) to 600px (~86 chars). A 200-char cell must clamp near the
    # NEW ceiling, well past the old one.
    w = _export_long_name_col_width(client)
    assert 50 < w <= 90, w


def test_autofit_cap_is_a_setting(client, monkeypatch):
    # U-2: DATA_ROVER_XLSX_AUTOFIT_MAX_PX tunes the cap (~px/7 char units).
    monkeypatch.setenv("DATA_ROVER_XLSX_AUTOFIT_MAX_PX", "150")
    w = _export_long_name_col_width(client)
    assert 0 < w < 25, w


def test_export_row_numbers_column(client):
    # Item 10: `show_row_numbers` prepends a 1-based "#" column, numbered in
    # export row order (which follows the current sort).
    _bootstrap_model(client)
    body = {
        "definition": {
            "row_source": {"kind": "scope", "types": ["Block"]},
            "show_row_numbers": True,
            "columns": [
                {"kind": "element", "source": {"kind": "row"}, "header": "Block"},
            ],
        }
    }
    r = client.post(papi("/tables/export"), json=body, headers=AUTH_HEADERS)
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws is not None
    header = [c.value for c in ws[1]]
    assert header[0] == "#"
    assert header[1] == "Block"
    numbers = [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value is not None]
    assert numbers == list(range(1, len(numbers) + 1))
    # the autofilter spans the "#" column too
    assert ws.auto_filter.ref.startswith("A1:B")


def test_export_row_numbers_off_by_default(client):
    _bootstrap_model(client)
    body = {
        "definition": {
            "row_source": {"kind": "scope", "types": ["Block"]},
            "columns": [
                {"kind": "element", "source": {"kind": "row"}, "header": "Block"},
            ],
        }
    }
    r = client.post(papi("/tables/export"), json=body, headers=AUTH_HEADERS)
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws is not None
    assert [c.value for c in ws[1]] == ["Block"]


def test_export_url_like_value_stays_a_plain_string(client):
    # A1 regression: xlsxwriter's default `strings_to_urls=True` routes any
    # string matching the url/mailto/file/(in|ex)ternal patterns through its
    # hyperlink writer instead of a plain string write. That writer silently
    # discards the cell past 65,530 URL cells/sheet or a 2079+ char URL (only
    # a `warnings.warn`, return code ignored) — a workbook shipped at HTTP
    # 200 with blank cells and no error. The `name` property is the only
    # string-typed property on Block in this metamodel, so it stands in for
    # "a property value" here.
    _bootstrap_model(client)
    client.post(
        papi("/model/elements"),
        json={
            "type": "Block",
            "properties": {"name": "https://example.com/x", "mass": 1.0},
        },
        headers=AUTH_HEADERS,
    )
    body = {
        "definition": {
            "row_source": {
                "kind": "scope",
                "types": ["Block"],
                "criteria": [
                    {
                        "type": "name_id",
                        "field": "name",
                        "op": "equals",
                        "value": "https://example.com/x",
                    }
                ],
            },
            "columns": [
                {
                    "kind": "property",
                    "source": {"kind": "row"},
                    "name": "name",
                    "header": "Name",
                },
            ],
        }
    }
    r = client.post(papi("/tables/export"), json=body, headers=AUTH_HEADERS)
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws is not None
    cell = ws["A2"]
    assert cell.value == "https://example.com/x"
    assert cell.hyperlink is None


def test_build_workbook_places_row_numbers_at_a_middle_column():
    # Regression guard for the reorderable row-number entry: the old API could
    # only PREPEND, so a row number anywhere but column A was unreachable.
    from data_rover.api.table_export import build_workbook
    from data_rover.core.metamodel.schema import Metamodel
    from data_rover.core.model.model import Model
    from data_rover.core.table.cells import ValueCell

    def v(text: str) -> ValueCell:
        return ValueCell(present=True, value=text, element_id=None, editable=False)

    blob = build_workbook(
        Model(Metamodel()),
        ["A", "#", "B"],
        "Sheet",
        [[v("a1"), v("b1")], [v("a2"), v("b2")]],
        row_number_col=1,
    )
    ws = load_workbook(io.BytesIO(blob)).active
    assert ws is not None
    assert [c.value for c in ws[1]] == ["A", "#", "B"]
    assert [c.value for c in ws[2]] == ["a1", 1, "b1"]
    assert [c.value for c in ws[3]] == ["a2", 2, "b2"]
    # the autofilter still spans every column, row number included
    assert ws.auto_filter.ref == "A1:C3"


def test_build_workbook_rejects_out_of_range_row_number_col():
    # Review finding: Task 4 will compute `row_number_col` dynamically from
    # user-draggable export settings, a caller that CAN get the index wrong.
    # Both too-high and negative must raise a clear `ValueError` (not a bare
    # `StopIteration`/`AssertionError`) so the route's
    # `except (NavigationResolveError, ValueError)` turns it into a 422.
    from data_rover.api.table_export import build_workbook
    from data_rover.core.metamodel.schema import Metamodel
    from data_rover.core.model.model import Model
    from data_rover.core.table.cells import ValueCell

    def v(text: str) -> ValueCell:
        return ValueCell(present=True, value=text, element_id=None, editable=False)

    for bad_col in (2, -1):
        with pytest.raises(ValueError, match="row_number_col"):
            build_workbook(
                Model(Metamodel()),
                ["A", "B"],
                "Sheet",
                [[v("a1"), v("b1")]],
                row_number_col=bad_col,
            )


def test_build_workbook_rejects_mismatched_row_length():
    # Companion to the out-of-range guard above: a row with the wrong cell
    # count (too few OR too many, relative to `len(headers)` minus the
    # row-number slot) must raise a clear `ValueError` rather than a bare
    # `StopIteration` or silently dropping extra cells.
    from data_rover.api.table_export import build_workbook
    from data_rover.core.metamodel.schema import Metamodel
    from data_rover.core.model.model import Model
    from data_rover.core.table.cells import ValueCell

    def v(text: str) -> ValueCell:
        return ValueCell(present=True, value=text, element_id=None, editable=False)

    # too few cells for a 3-header, 1-row-number-column sheet (expects 2)
    with pytest.raises(ValueError, match="row 1 has 1 cell"):
        build_workbook(
            Model(Metamodel()),
            ["A", "#", "B"],
            "Sheet",
            [[v("a1")]],
            row_number_col=1,
        )
    # too many cells for the same sheet
    with pytest.raises(ValueError, match="row 1 has 3 cell"):
        build_workbook(
            Model(Metamodel()),
            ["A", "#", "B"],
            "Sheet",
            [[v("a1"), v("b1"), v("extra")]],
            row_number_col=1,
        )


def _two_col_body(**defn_over):
    defn = {
        "row_source": {"kind": "scope", "types": ["Block"]},
        "columns": [
            {"kind": "element", "source": {"kind": "row"}, "header": "Block"},
            {
                "kind": "property",
                "source": {"kind": "row"},
                "name": "mass",
                "header": "Mass",
            },
        ],
    }
    defn.update(defn_over)
    return {"definition": defn}


def test_export_xlsx_honors_export_order_and_header_override(client):
    _bootstrap_model(client)
    body = _two_col_body(export_order=[1, 0])
    body["definition"]["columns"][0]["export"] = {"header": "Assembly"}
    r = client.post(papi("/tables/export"), json=body, headers=AUTH_HEADERS)
    assert r.status_code == 200
    ws = load_workbook(io.BytesIO(r.content)).active
    assert ws is not None
    assert [c.value for c in ws[1]] == ["Mass", "Assembly"]


def test_export_xlsx_excludes_an_opted_out_column(client):
    _bootstrap_model(client)
    body = _two_col_body()
    body["definition"]["columns"][1]["export"] = {"include": False}
    r = client.post(papi("/tables/export"), json=body, headers=AUTH_HEADERS)
    ws = load_workbook(io.BytesIO(r.content)).active
    assert ws is not None
    assert [c.value for c in ws[1]] == ["Block"]


def test_export_xlsx_includes_an_opted_in_hidden_column(client):
    _bootstrap_model(client)
    body = _two_col_body()
    body["definition"]["columns"][1]["hidden"] = True
    body["definition"]["columns"][1]["export"] = {"include": True}
    r = client.post(papi("/tables/export"), json=body, headers=AUTH_HEADERS)
    ws = load_workbook(io.BytesIO(r.content)).active
    assert ws is not None
    assert [c.value for c in ws[1]] == ["Block", "Mass"]


def test_export_xlsx_row_number_column_can_be_moved_and_renamed(client):
    _bootstrap_model(client)
    body = _two_col_body(
        show_row_numbers=True,
        export_order=[0, -1, 1],
        export_row_number={"header": "No."},
    )
    r = client.post(papi("/tables/export"), json=body, headers=AUTH_HEADERS)
    ws = load_workbook(io.BytesIO(r.content)).active
    assert ws is not None
    assert [c.value for c in ws[1]] == ["Block", "No.", "Mass"]
    assert ws.cell(row=2, column=2).value == 1


def test_export_xlsx_excluded_column_before_the_row_number_column(client):
    # The combination `test_export_xlsx_row_number_column_can_be_moved_and_
    # renamed` does not reach: the row-number slot sits AFTER a column the
    # export drops, so its output position only lines up if `export_layout`
    # skips the dropped column without spending a position on it. A stale
    # position would put "#" on the wrong column here, or (one past the end)
    # trip `build_workbook`'s ValueError guard and 422 the whole export.
    _bootstrap_model(client)
    body = _two_col_body(show_row_numbers=True, export_order=[0, -1, 1])
    body["definition"]["columns"][0]["export"] = {"include": False}
    r = client.post(papi("/tables/export"), json=body, headers=AUTH_HEADERS)
    assert r.status_code == 200, r.text
    ws = load_workbook(io.BytesIO(r.content)).active
    assert ws is not None
    assert [c.value for c in ws[1]] == ["#", "Mass"]
    assert ws.cell(row=2, column=1).value == 1
    # the autofilter still spans both output columns
    assert ws.auto_filter.ref.startswith("A1:B")


def test_export_xlsx_row_number_column_can_be_excluded(client):
    _bootstrap_model(client)
    body = _two_col_body(show_row_numbers=True, export_row_number={"include": False})
    r = client.post(papi("/tables/export"), json=body, headers=AUTH_HEADERS)
    ws = load_workbook(io.BytesIO(r.content)).active
    assert ws is not None
    assert [c.value for c in ws[1]] == ["Block", "Mass"]


def test_export_xlsx_unchanged_without_export_settings(client):
    # The no-migration guarantee, asserted rather than assumed: a definition
    # that carries no export settings must produce the same sheet it did
    # before this feature existed. Structural, not a byte comparison — xlsx
    # bytes are not reproducible across runs.
    _bootstrap_model(client)
    body = _two_col_body(show_row_numbers=True)
    first = client.post(papi("/tables/export"), json=body, headers=AUTH_HEADERS)
    ws = load_workbook(io.BytesIO(first.content)).active
    assert ws is not None
    assert [c.value for c in ws[1]] == ["#", "Block", "Mass"]
    assert ws.cell(row=2, column=1).value == 1


class TestSheetTitle:
    """Direct unit tests for `_sheet_title`; it had zero before this fix."""

    def test_forbidden_chars_replaced(self):
        assert _sheet_title("a[b]:c*d?e/f\\g") == "a_b__c_d_e_f_g"

    def test_leading_trailing_apostrophe_stripped(self):
        assert _sheet_title("'quoted name'") == "quoted name"

    def test_empty_name_falls_back_to_table(self):
        assert _sheet_title("") == "Table"

    def test_all_apostrophes_name_falls_back_to_table(self):
        # The only way `cleaned.strip("'")` (unchanged by this fix, see
        # module docstring / A3 report note) produces an empty string from a
        # non-empty input: a name made entirely of apostrophes/quotes.
        assert _sheet_title("''''") == "Table"

    def test_31_char_name_with_apostrophe_as_32nd_char_regression(self):
        # A3 regression: stripping BEFORE truncating would keep this name at
        # 31 chars ending in "'" (the truncation would land exactly after
        # the leading 31 chars, all non-apostrophe, but the ORIGINAL bug was
        # stripping outer quotes first and truncating second — reproduced
        # here as a name whose 31st character is itself an apostrophe, which
        # must survive truncation and then be stripped).
        name = "a" * 30 + "'" + "extra text past the limit"
        assert len(name) > 31
        assert name[30] == "'"
        title = _sheet_title(name)
        assert not title.endswith("'")
        assert title == "a" * 30

    def test_plain_truncation_to_31_chars(self):
        name = "x" * 40
        title = _sheet_title(name)
        assert title == "x" * 31
        assert len(title) == 31
