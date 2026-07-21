"""Evaluate-route Phase B behaviour (spec 2026-07-20 §4.1-4.2).

The whole-table passes (`build_rows_ex` + `order_rows`) run CACHE-ONLY: the
guest is never driven O(rows) times inside a request. A miss records a pending
cell, which makes the route degrade to BUILD order (a sort over half-pending
values would visibly reshuffle on every poll), kick/join the background sweep,
and report a `script_status` block the client polls on. Only the visible window
is still evaluated live.

Fixtures mirror `test_script_cell_cache_api.py` (separate `app`/`client` so the
runner can be swapped through `dependency_overrides`); the runner fakes come
from `_script_fakes.py`. Sync sweep mode is pinned exactly the way
`test_script_sweep.py` pins it — the `DATA_ROVER_SNIPPET_SWEEP_SYNC` env var
plus a fresh `get_settings()` and an assertion that the flag took — so the
"after the sweep" cases observe a finished job with no sleeping. The one async
case parks the sweep thread on an Event instead.
"""

from __future__ import annotations

import hashlib
import io
import threading
from collections.abc import Iterator

import httpx
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from data_rover.api.main import create_app
from data_rover.api.routes.tables import _status_from_job
from data_rover.api.script_runner import get_runner
from data_rover.api.script_sweep import SweepJob
from data_rover.api.session import get_session
from data_rover.api.settings import Settings, get_settings
from data_rover.api.table_cache import table_fingerprint
from data_rover.core.script.runner import CallResult
from data_rover.core.table.schema import TABLE_ADAPTER

from ._script_fakes import CountingRunner, ScriptedRunner, ok, timeout
from .conftest import AUTH_HEADERS, papi, seed_default_project

THING_MM = """
elements:
  - name: Thing
    key: [name]
    properties:
      - {name: name, datatype: string, multiplicity: "1"}
"""

VALUE_CODE = "def value(els): return 1"

#: Five elements: enough that the sweep's consecutive-timeout abort threshold
#: (3) and a two-row window produce distinguishable guest-call counts.
THING_IDS = ["t1", "t2", "t3", "t4", "t5"]


@pytest.fixture
def app() -> Iterator[FastAPI]:
    seed_default_project()
    application = create_app()
    yield application
    application.dependency_overrides.clear()


@pytest.fixture
def client(app: FastAPI) -> TestClient:
    c = TestClient(app)
    c.headers.update(AUTH_HEADERS)
    return c


@pytest.fixture
def seed_thing_model(client: TestClient) -> None:
    """`Thing` metamodel + five `Thing` elements, loaded through the HTTP
    routes so the table below has real rows in a deterministic build order
    (`t1`..`t5`, the scope's insertion order)."""
    r = client.post(
        papi("/metamodel"),
        content=THING_MM,
        headers={"content-type": "application/x-yaml"},
    )
    assert r.status_code == 200, r.text
    r = client.post(
        papi("/model"),
        json={
            "elements": [
                {"id": tid, "type_name": "Thing", "properties": {"name": f"N{i}"}}
                for i, tid in enumerate(THING_IDS)
            ],
            "relationships": [],
        },
    )
    assert r.status_code == 200, r.text


@pytest.fixture
def settings_sync_sweep(monkeypatch: pytest.MonkeyPatch) -> Settings:
    """Pin `snippet_sweep_sync=True` (env var + fresh `get_settings()`), so
    `kick_or_join_sweep` runs the whole sweep inline inside the request that
    kicks it. `get_settings` is an uncached `Settings()` factory, so the route's
    own `Depends(get_settings)` picks the env var up too."""
    monkeypatch.setenv("DATA_ROVER_SNIPPET_SWEEP_SYNC", "true")
    settings = get_settings()
    assert settings.snippet_sweep_sync is True
    return settings


def _script_table() -> dict:
    """Element column + a COLLAPSE script column. `keep_empty` defaults True, so
    an UNSORTED evaluation never calls the snippet during the whole-table
    passes — only a sort on column 1 does."""
    return {
        "row_source": {"kind": "scope", "types": ["Thing"]},
        "columns": [
            {"kind": "element"},
            {
                "kind": "script",
                "snippet": {"definition": {"code": VALUE_CODE}},
            },
        ],
    }


#: An EXPAND script column whose snippet returns a None scalar: `_expand_values`
#: promotes no binding, `keep_empty` (default True) keeps one row per element
#: with a None key slot, and `cells.py` re-derives that slot's cell with a
#: FORCED cache-only call. That forced call is the only production path that can
#: record a pending miss on an ORDER-CACHE HIT (there is no per-request memo to
#: serve it), which is exactly what FIX 2's path (a) is about.
EMPTY_EXPAND_CODE = "def value(els): return None"


def _expand_script_table() -> dict:
    return {
        "row_source": {"kind": "scope", "types": ["Thing"]},
        "columns": [
            {"kind": "element"},
            {
                "kind": "script",
                "mode": "expand",
                "snippet": {"definition": {"code": EMPTY_EXPAND_CODE}},
            },
        ],
    }


def _plain_table() -> dict:
    """No script column anywhere: `script_ctx is None`, so the route must be
    completely untouched by Phase B (`script_status is None`)."""
    return {
        "row_source": {"kind": "scope", "types": ["Thing"]},
        "columns": [{"kind": "element"}],
    }


def _fingerprint(table: dict | None = None) -> str:
    """The sweep's job key for a table: the resolved definition dumped with a
    None sort (the job key excludes the sort on purpose)."""
    defn = TABLE_ADAPTER.validate_python(table if table is not None else _script_table())
    return table_fingerprint(TABLE_ADAPTER.dump_json(defn).decode(), None)


def _descending_value(_i: int, ids: list[str]) -> CallResult:
    """`t1`->9 … `t5`->5, so an ASCENDING sort on the script column yields the
    exact REVERSE of the build order — build order and sorted order can never
    be confused for one another."""
    return ok(10 - int(ids[0][1:]))


def _evaluate(
    client: TestClient, table: dict, *, sort: bool = True, limit: int = 50
) -> dict:
    body: dict = {"definition": table, "limit": limit}
    if sort:
        body["sort"] = {"column": 1, "direction": "asc"}
    r = client.post(papi("/tables/evaluate"), json=body, headers=AUTH_HEADERS)
    assert r.status_code == 200, r.text
    return r.json()


def _keys(page: dict) -> list[object]:
    return [row["key"][0] for row in page["rows"]]


def _export(
    client: TestClient, table: dict, *, sort: bool = False
) -> httpx.Response:
    body: dict = {"definition": table}
    if sort:
        body["sort"] = {"column": 1, "direction": "asc"}
    return client.post(papi("/tables/export"), json=body, headers=AUTH_HEADERS)


def _sheet(resp: httpx.Response) -> list[list[object]]:
    """The exported sheet as a list of rows (header row included, and — when
    the export degraded — the trailing single-cell notice row)."""
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws is not None
    return [[c.value for c in row] for row in ws.iter_rows()]


def _data_rows(resp: httpx.Response) -> list[list[object]]:
    """Just the `THING_IDS` rows: header stripped off the front and any notice
    row off the back, so an assertion on cell values is unaffected by whether
    the export flagged itself as degraded."""
    return _sheet(resp)[1 : 1 + len(THING_IDS)]


# --------------------------------------------------------------------------
# _status_from_job: the job-state -> wire-state mapping
# --------------------------------------------------------------------------


def test_status_from_job_never_reports_ready() -> None:
    """D1: `_status_from_job` is only reached when the cache-only pass saw
    pending, so its rows predate the sweep — even a job that FINISHED during
    this very request (sync mode) must report `computing` so the client polls
    once more and gets the clean sorted page."""
    running = _status_from_job(SweepJob(fingerprint="f", rev=0, done=1, total=4))
    assert running.state == "computing"
    assert (running.done, running.total) == (1, 4)

    finished = SweepJob(fingerprint="f", rev=0, state="done", done=4, total=4)
    assert _status_from_job(finished).state == "computing"


def test_status_from_job_maps_failed_and_cancelled_to_terminal_states() -> None:
    """`failed` and `cancelled` are both DEAD jobs — no thread is behind either
    — so both map to the terminal wire state `failed`. Reporting `computing`
    for a cancelled job would strand the poller until the next commit."""
    failed = SweepJob(
        fingerprint="f", rev=0, state="failed", message="boom", done=2, total=9
    )
    out = _status_from_job(failed)
    assert out.state == "failed"
    assert out.message == "boom"
    assert (out.done, out.total) == (2, 9)

    cancelled = SweepJob(
        fingerprint="f", rev=0, state="cancelled", message="sweep cancelled"
    )
    out = _status_from_job(cancelled)
    assert out.state == "failed"
    assert out.message == "sweep cancelled"


# --------------------------------------------------------------------------
# Route behaviour
# --------------------------------------------------------------------------


def test_plain_table_has_no_script_status(
    client: TestClient, app: FastAPI, seed_thing_model: None
) -> None:
    """Regression net: a table without script columns is untouched."""
    page = _evaluate(client, _plain_table(), sort=False)
    assert page["script_status"] is None
    assert _keys(page) == THING_IDS


def test_unsorted_default_table_stays_inline(
    client: TestClient, app: FastAPI, seed_thing_model: None
) -> None:
    """collapse + keep_empty + no sort: the whole-table passes never touch the
    snippet, so nothing is pending, no sweep is registered, and the window's
    cells carry real values on the FIRST response."""
    counting = CountingRunner()
    app.dependency_overrides[get_runner] = lambda: counting

    page = _evaluate(client, _script_table(), sort=False)
    assert page["script_status"] == {
        "state": "ready",
        "done": 0,
        "total": None,
        "message": None,
    }
    assert _keys(page) == THING_IDS
    assert [row["cells"][1]["kind"] for row in page["rows"]] == ["value"] * 5
    assert counting.calls == 5  # the visible window really was computed live
    session = get_session()
    assert session.script_sweeps.get(_fingerprint(), session.model_rev) is None


def test_partially_cached_sort_degrades_to_build_order(
    client: TestClient,
    app: FastAPI,
    seed_thing_model: None,
    settings_sync_sweep: Settings,
) -> None:
    """THE discriminating test for "degrade to build order".

    A FULLY cold table cannot prove it: `_sort_value` maps every pending result
    onto the same `(1, ())` empty key, so sorting an all-pending table is a
    stable no-op that already equals build order. The failure the degrade
    actually prevents is a PARTIALLY cached table visibly reshuffling on every
    poll — so this test warms the cell cache for a strict SUBSET of the rows.

    `t4`/`t5` are pre-seeded with 6/5 (the values the runner would produce);
    `t1`..`t3` stay pending. Ascending sort on the script column would put the
    two non-empty rows FIRST (empties always sort last, in both directions),
    yielding `[t5, t4, t1, t2, t3]` — distinct from BOTH the build order
    `[t1..t5]` this route must return AND the fully-warm sort
    `[t5, t4, t3, t2, t1]` a later poll returns. Three mutually
    distinguishable orders, so the assertion below can only pass for one of
    them.
    """
    runner = ScriptedRunner(_descending_value)
    app.dependency_overrides[get_runner] = lambda: runner

    session = get_session()
    sha = hashlib.sha256(VALUE_CODE.encode()).hexdigest()
    for tid in ("t4", "t5"):
        session.script_cell_cache.put(
            (sha, "value", (tid,)), ok(10 - int(tid[1:])), session.model_rev
        )

    page = _evaluate(client, _script_table())
    assert page["script_status"]["state"] == "computing"
    assert _keys(page) == THING_IDS  # NOT ["t5","t4","t1","t2","t3"]

    # ...and once the (sync) sweep has filled the rest, the sort is the clean
    # fully-warm one — proving the half-sorted interleave was never the answer.
    second = _evaluate(client, _script_table())
    assert second["script_status"]["state"] == "ready"
    assert _keys(second) == list(reversed(THING_IDS))


def test_sorted_script_table_while_computing(
    client: TestClient, app: FastAPI, seed_thing_model: None
) -> None:
    """ASYNC sweep parked on an Event: the request must still answer promptly
    with `computing` + BUILD-order rows, while the window itself renders live
    values.

    The fake blocks calls made from the sweep's own daemon thread (named
    `script-sweep`) and answers every other call immediately — that is the
    deterministic sync point, no sleeping and no self-deadlock when the route
    evaluates the visible window with the very same runner object.

    `limit=2` removes the scheduling race the first cut of this test had. The
    sweep and the window compute the SAME `(code, "value", ids)` keys at the
    same rev, so with a full-width window the sweep could get nothing but cache
    hits, never call the runner, and never set `entered`. A two-row window
    leaves `t3`..`t5` computable ONLY by the sweep, so it must reach the
    blocking fake whatever the interleaving.
    """
    entered = threading.Event()
    proceed = threading.Event()
    sweep_thread: list[threading.Thread] = []

    def _outcome(i: int, ids: list[str]) -> CallResult:
        if threading.current_thread().name == "script-sweep":
            sweep_thread.append(threading.current_thread())
            entered.set()
            proceed.wait(timeout=10.0)
        return _descending_value(i, ids)

    runner = ScriptedRunner(_outcome)
    app.dependency_overrides[get_runner] = lambda: runner

    try:
        page = _evaluate(client, _script_table(), limit=2)
        assert page["script_status"]["state"] == "computing"
        # Degraded: build order, NOT the requested descending-value sort.
        assert _keys(page) == THING_IDS[:2]
        # The visible window is still evaluated live, so its cells are real.
        assert [row["cells"][1]["kind"] for row in page["rows"]] == ["value"] * 2
        # The sweep thread really is running behind that `computing`.
        assert entered.wait(timeout=5.0)
    finally:
        proceed.set()  # release the parked call so the daemon thread exits
        # Deterministic teardown: the released daemon thread keeps writing into
        # this session's cell cache, so JOIN it rather than letting the next
        # test race it.
        if sweep_thread:
            sweep_thread[0].join(timeout=10.0)
            assert not sweep_thread[0].is_alive()


def test_sorted_script_table_after_sweep(
    client: TestClient,
    app: FastAPI,
    seed_thing_model: None,
    settings_sync_sweep: Settings,
) -> None:
    """SYNC sweep: the first response is deliberately asymmetric.

    `kick_or_join_sweep` runs the entire sweep inside the first request, so the
    job is already `done` when the route samples it — but that request's rows
    were built by the cache-only pass BEFORE those values existed, so it reports
    `computing` (never `ready`) and the client polls once more. The SECOND
    request finds every value cached, sorts cleanly, reports `ready`, and
    populates the order cache.
    """
    runner = ScriptedRunner(_descending_value)
    app.dependency_overrides[get_runner] = lambda: runner

    first = _evaluate(client, _script_table())
    status = first["script_status"]
    assert status["state"] == "computing"
    assert status["done"] == status["total"] == 5  # the sync sweep finished
    assert _keys(first) == THING_IDS  # ...but these rows are still build order

    second = _evaluate(client, _script_table())
    assert second["script_status"]["state"] == "ready"
    assert _keys(second) == list(reversed(THING_IDS))  # sorted by script value
    assert [row["cells"][1]["value"] for row in second["rows"]] == [5, 6, 7, 8, 9]

    # Third call with a pristine runner: the order cache serves the sort and the
    # cell cache serves the cells, so the guest is not touched at all.
    fresh = CountingRunner()
    app.dependency_overrides[get_runner] = lambda: fresh
    third = _evaluate(client, _script_table())
    assert third["script_status"]["state"] == "ready"
    assert _keys(third) == list(reversed(THING_IDS))
    assert fresh.calls == 0


def test_failed_sweep_reported_not_rekicked(
    client: TestClient,
    app: FastAPI,
    seed_thing_model: None,
    settings_sync_sweep: Settings,
) -> None:
    """A sweep aborted by the consecutive-timeout guard surfaces as
    `state="failed"` with the job's message, and failed-job memory keeps the
    next poll from restarting the grind."""
    runner = ScriptedRunner(lambda i, ids: timeout())
    app.dependency_overrides[get_runner] = lambda: runner
    abort_at = settings_sync_sweep.snippet_sweep_timeout_abort

    first = _evaluate(client, _script_table(), limit=2)
    assert first["script_status"]["state"] == "failed"
    assert "consecutive" in first["script_status"]["message"]
    assert _keys(first) == THING_IDS[:2]
    # sweep aborted at the threshold + the two live window cells
    assert runner.calls[0] == abort_at + 2
    after_first = runner.calls[0]

    second = _evaluate(client, _script_table(), limit=2)
    assert second["script_status"]["state"] == "failed"
    # Only the two live window cells ran again: no second sweep (which would
    # have added another `abort_at` calls).
    assert runner.calls[0] == after_first + 2


def test_order_cache_hit_with_evicted_cell_downgrades_to_computing(
    client: TestClient,
    app: FastAPI,
    seed_thing_model: None,
    settings_sync_sweep: Settings,
) -> None:
    """FIX 2, path (a): an order-cache HIT must not report `ready` when the
    window itself goes pending.

    The order cache and the cell cache are bounded INDEPENDENTLY, so a warm
    order can outlive the cell entries it was built from. An `expand` script
    column re-derives its cell with a FORCED cache-only call (cells.py) and on
    an order-cache hit there is no per-request memo to serve it — so an evicted
    cell entry yields a `PendingCell`. Reporting `ready` there would stop the
    client polling and strand that cell until the rev moves.
    """
    runner = ScriptedRunner(lambda i, ids: ok(None))
    app.dependency_overrides[get_runner] = lambda: runner
    table = _expand_script_table()

    # 1st: cold — pending everywhere, sync sweep fills the cell cache.
    assert _evaluate(client, table, sort=False)["script_status"]["state"] == "computing"
    # 2nd: fully warm — `ready`, and THIS is the response that stores the order.
    assert _evaluate(client, table, sort=False)["script_status"]["state"] == "ready"

    session = get_session()
    fp = _fingerprint(table)
    rev = session.model_rev
    assert session.table_order_cache.get(fp, "none", rev) is not None
    # Simulate the independent LRU evictions: drop the cell entries (keeping the
    # rev stamp, so writes still land) and forget the finished sweep job.
    session.script_cell_cache.clear_and_stamp(rev)
    session.script_sweeps.cancel_all()
    assert session.script_sweeps.get(fp, rev) is None

    page = _evaluate(client, table, sort=False)
    assert session.table_order_cache.get(fp, "none", rev) is not None  # still a HIT
    assert [row["cells"][1]["kind"] for row in page["rows"]] == ["pending"] * 5
    assert page["script_status"]["state"] == "computing"  # NOT "ready"
    assert session.script_sweeps.get(fp, rev) is not None  # ...and a sweep was kicked


def test_runner_unavailable_reports_failed_without_a_sweep(
    client: TestClient, app: FastAPI, seed_thing_model: None
) -> None:
    """No runner at all: there is nothing to sweep with, so the route reports
    `failed` immediately instead of kicking a job that could only fail."""
    app.dependency_overrides[get_runner] = lambda: None

    page = _evaluate(client, _script_table())
    assert page["script_status"] == {
        "state": "failed",
        "done": 0,
        "total": None,
        "message": "script runner unavailable",
    }
    assert _keys(page) == THING_IDS
    session = get_session()
    assert session.script_sweeps.get(_fingerprint(), session.model_rev) is None


# --------------------------------------------------------------------------
# /tables/export: 202 while the sweep is computing (Task 8, spec §4.4)
# --------------------------------------------------------------------------


def test_export_plain_table_never_202s(
    client: TestClient, app: FastAPI, seed_thing_model: None
) -> None:
    """Regression net: no script column => `script_ctx is None` => no probe, no
    202, byte-for-byte the pre-Phase-B export."""
    app.dependency_overrides[get_runner] = lambda: CountingRunner()
    r = _export(client, _plain_table())
    assert r.status_code == 200
    assert len(_sheet(r)) == 1 + len(THING_IDS)


def test_export_display_only_script_column_202s_on_cold_cache(
    client: TestClient,
    app: FastAPI,
    seed_thing_model: None,
    settings_sync_sweep: Settings,
) -> None:
    """THE completeness-probe test.

    `_script_table()`'s script column is a plain COLLAPSE `keep_empty=True`
    DISPLAY column with no sort on it: `build_rows`/`order_rows` never invoke
    its `value()` at all, so a probe made of those two passes alone reports
    `pending_misses == 0` on a stone-cold cache and the export 200s — full of
    silent `#ERROR: not computed` cells. Only the extra whole-table
    `evaluate_cells` pass sees the display column, so this must be a 202.
    """
    runner = ScriptedRunner(_descending_value)
    app.dependency_overrides[get_runner] = lambda: runner

    r = _export(client, _script_table())
    assert r.status_code == 202, r.text
    assert r.headers["retry-after"] == "1"
    assert r.json()["state"] == "computing"

    # ...and the retry (the sync sweep already filled the cache) ships values.
    second = _export(client, _script_table())
    assert second.status_code == 200
    assert [row[1] for row in _sheet(second)[1:]] == [9, 8, 7, 6, 5]


def test_export_202_while_computing_then_200(
    client: TestClient,
    app: FastAPI,
    seed_thing_model: None,
    settings_sync_sweep: Settings,
) -> None:
    """Sorted export: 202 + `Retry-After: 1` + a `computing` body first, then a
    200 xlsx carrying the sorted real values — served entirely from the cell
    cache, with a pristine runner proving NO fresh guest call was made."""
    runner = ScriptedRunner(_descending_value)
    app.dependency_overrides[get_runner] = lambda: runner

    first = _export(client, _script_table(), sort=True)
    assert first.status_code == 202, first.text
    assert first.headers["retry-after"] == "1"
    body = first.json()
    assert body["state"] == "computing"
    assert body["done"] == body["total"] == len(THING_IDS)  # the sync sweep ran
    assert first.headers["content-type"].startswith("application/json")

    fresh = CountingRunner()
    app.dependency_overrides[get_runner] = lambda: fresh
    second = _export(client, _script_table(), sort=True)
    assert second.status_code == 200, second.text
    assert second.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    rows = _sheet(second)[1:]
    # ascending by script value => t5..t1 => 5,6,7,8,9
    assert [row[1] for row in rows] == [5, 6, 7, 8, 9]
    assert "x-table-script-errors" not in second.headers
    assert fresh.calls == 0  # cache hits only: no guest work during a download


def test_export_failed_sweep_ships_error_cells(
    client: TestClient,
    app: FastAPI,
    seed_thing_model: None,
    settings_sync_sweep: Settings,
) -> None:
    """A sweep aborted by the consecutive-timeout guard is TERMINAL: promising
    the client `computing` would make it retry forever, so the export falls
    through and ships the honest answer — pending rendered `#ERROR`.

    Timeouts are never cached and a cache-only miss records a PENDING cell (not
    an error), so `script_ctx.errored` stays False — but the cells still SHIP as
    `#ERROR`, so the route must not stay silent: it flags the workbook with
    `X-Table-Script-Errors` and a trailing notice row on the strength of the
    pending misses the render itself recorded (FIX 2)."""
    runner = ScriptedRunner(lambda i, ids: timeout())
    app.dependency_overrides[get_runner] = lambda: runner

    r = _export(client, _script_table())
    assert r.status_code == 200, r.text
    assert [row[1] for row in _data_rows(r)] == ["#ERROR: not computed"] * len(
        THING_IDS
    )
    assert r.headers["x-table-script-errors"] == "true"
    notice = _sheet(r)[-1][0]
    assert isinstance(notice, str) and "#ERROR" in notice

    # Failed-job memory: the retry is served the SAME dead job, so no second
    # grind is started (only the abort threshold's calls were ever made).
    calls_after_first = runner.calls[0]
    assert calls_after_first == settings_sync_sweep.snippet_sweep_timeout_abort
    again = _export(client, _script_table())
    assert again.status_code == 200
    assert runner.calls[0] == calls_after_first


def test_export_cancelled_sweep_does_not_202_forever(
    client: TestClient, app: FastAPI, seed_thing_model: None
) -> None:
    """A `cancelled` job has no thread behind it either. `kick()` hands the same
    dead job back at the same rev, so answering 202 would strand the client in
    an endless retry loop — the route reads the wire state from
    `_status_from_job` (which collapses `cancelled` onto `failed`) and exports.
    """
    runner = ScriptedRunner(_descending_value)
    app.dependency_overrides[get_runner] = lambda: runner
    session = get_session()

    def _start(job: SweepJob) -> None:
        job.state = "cancelled"
        job.message = "sweep cancelled"

    job = session.script_sweeps.kick(_fingerprint(), session.model_rev, _start)
    assert job.state == "cancelled"

    r = _export(client, _script_table())
    assert r.status_code == 200, r.text
    assert [row[1] for row in _data_rows(r)] == ["#ERROR: not computed"] * len(
        THING_IDS
    )
    # No fresh job was started behind the cancelled one.
    assert session.script_sweeps.get(_fingerprint(), session.model_rev) is job


def test_export_done_sweep_with_permanent_hole_does_not_202_forever(
    client: TestClient,
    app: FastAPI,
    seed_thing_model: None,
    settings_sync_sweep: Settings,
) -> None:
    """FIX 1 — the ENDLESS-202 regression.

    A `done` sweep does NOT imply a complete cache. `ScriptCellCache.put`
    refuses non-deterministic error kinds (`timeout`/`unavailable`/`cancelled`)
    while the sweep only aborts on a run of `snippet_sweep_timeout_abort` (3)
    CONSECUTIVE ones — so a single intermittently-timing-out cell (here: guest
    call #0, i.e. `t1`) leaves a permanent hole behind an otherwise successful
    sweep. Deciding 202 off "the job is not failed" alone would then loop
    forever: the same rev hands back the same `done` job, `_status_from_job`
    maps `done` -> `computing`, and the file becomes undownloadable.

    The route instead RE-PROBES the cache after joining the sweep: terminal job
    + still-pending cells => nothing will ever fill them at this rev => ship the
    honest terminal export. So the very first call is a 200, and it stays a 200
    however many times the client asks."""
    runner = ScriptedRunner(lambda i, ids: timeout() if i == 0 else ok(7))
    app.dependency_overrides[get_runner] = lambda: runner

    responses: list[httpx.Response] = []
    for attempt in range(5):
        r = _export(client, _script_table())
        assert r.status_code == 200, f"attempt {attempt}: {r.status_code} {r.text}"
        responses.append(r)

    last = responses[-1]
    rows = _data_rows(last)
    # `t1` was the timed-out call: never cached, so it ships as `#ERROR`; the
    # other four were swept successfully and carry real values.
    assert [row[1] for row in rows] == ["#ERROR: not computed"] + [7] * 4
    # ...and the degraded workbook says so (FIX 2).
    assert last.headers["x-table-script-errors"] == "true"
    # Exactly one sweep ran (5 guest calls): failed-job memory kept the retries
    # from re-grinding, which is also why the hole is permanent at this rev.
    assert runner.calls[0] == len(THING_IDS)


def test_export_without_runner_flags_error_cells(
    client: TestClient, app: FastAPI, seed_thing_model: None
) -> None:
    """FIX 3 — the runner-unavailable export path.

    With no runner there is nothing to sweep with, so the route never kicks a
    job and falls straight through. Its context is cache-only like every other
    export context, so the cells come back `pending` (rendered
    `#ERROR: not computed`) rather than `unavailable` — which leaves
    `script_ctx.errored` False. The signal must still be there: header plus
    trailing notice row, so nobody mistakes an all-`#ERROR` workbook for an
    authoritative one."""
    app.dependency_overrides[get_runner] = lambda: None

    r = _export(client, _script_table())
    assert r.status_code == 200, r.text
    assert [row[1] for row in _data_rows(r)] == ["#ERROR: not computed"] * len(
        THING_IDS
    )
    assert r.headers["x-table-script-errors"] == "true"
    notice = _sheet(r)[-1][0]
    assert isinstance(notice, str) and "#ERROR" in notice
    # No runner => no sweep was kicked (there would be nothing to sweep with).
    session = get_session()
    assert session.script_sweeps.get(_fingerprint(), session.model_rev) is None
