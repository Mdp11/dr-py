"""xlsx writer for table export. Lives in the API layer (core stays xlsx-free).
Consumes core cell dataclasses and produces workbook bytes.

`write_only` mode + `WriteOnlyCell` streams rows straight to the zip-backed
workbook writer rather than building an in-memory worksheet object graph —
the whole point of export being uncapped (see routes/tables.py) is that a
table can be much larger than a page, so keeping peak memory proportional to
one chunk (not the whole sheet) matters here.
"""

from __future__ import annotations

import io
from collections.abc import Callable, Iterable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from data_rover.core.model.model import Model
from data_rover.core.model.naming import display_name
from data_rover.core.table.cells import (
    Cell,
    ElementCell,
    ElementsCell,
    ErrorCell,
    ValueCell,
    ValuesCell,
)


def _display(model: Model, eid: str) -> str:
    # shared case-insensitive `name` lookup — same label the grid displays
    return display_name(model.elements[eid])


def _cell_text(model: Model, cell: Cell) -> object:
    """Map one core cell dataclass to the xlsx value it should render as."""
    if isinstance(cell, ElementCell):
        return _display(model, cell.element_id) if cell.element_id else ""
    if isinstance(cell, ValueCell):
        return "" if not cell.present or cell.value is None else cell.value
    if isinstance(cell, ValuesCell):
        return "; ".join(str(v) for v in cell.values)
    if isinstance(cell, ErrorCell):
        return f"#ERROR: {cell.message}"
    assert isinstance(cell, ElementsCell)
    return "; ".join(_display(model, e) for e in cell.element_ids)


def build_workbook(
    model: Model,
    headers: list[str],
    widths: list[int | None],
    sheet_name: str,
    row_iter: Iterable[list[Cell]],
    *,
    notice_provider: Callable[[], str | None] | None = None,
) -> bytes:
    """Stream `row_iter` into a single-sheet workbook: bold+frozen header row,
    heuristic column widths from `width_px` (`max(4, w / 7)` — openpyxl widths
    are in roughly-character units, not px), sheet name truncated to xlsx's
    31-char limit. Returns the finished workbook's bytes.

    `notice_provider`, if given, is called AFTER `row_iter` is fully consumed
    (not before) — callers whose "should there be a notice" flag only settles
    once every lazily-evaluated cell has been visited (e.g. a script column's
    error flag) must defer that decision to this point rather than computing
    it up front. A truthy return appends one trailing single-cell row."""
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=(sheet_name or "Table")[:31])

    header_cells = []
    for h in headers:
        c = WriteOnlyCell(ws, value=h)
        c.font = Font(bold=True)
        header_cells.append(c)
    ws.append(header_cells)
    ws.freeze_panes = "A2"
    for i, w in enumerate(widths):
        if w:
            ws.column_dimensions[get_column_letter(i + 1)].width = max(4, w / 7)
    for row in row_iter:
        ws.append([_cell_text(model, c) for c in row])

    if notice_provider is not None:
        text = notice_provider()
        if text:
            ws.append([WriteOnlyCell(ws, value=text)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
